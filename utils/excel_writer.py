"""
Excel 凭证输出工具
将凭证行列表写入符合金蝶/用友导入格式的 xlsx，黄色行标注。
"""
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

OUTPUT_COLS = [
    '日期', '凭证字', '凭证号', '附件数', '分录序号', '摘要',
    '科目代码', '科目名称',
    '借方金额', '贷方金额',
    '客户', '供应商', '职员', '项目', '部门', '存货',
    '自定义辅助核算类别', '自定义辅助核算编码',
    '自定义辅助核算类别1', '自定义辅助核算编码1',
    '数量', '单价', '原币金额', '币别', '汇率',
]

YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
HEADER_FONT = Font(bold=True)
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_voucher_excel(rows: list[dict], sheet_name: str = '凭证') -> BytesIO:
    """
    rows: list of dicts，每个 dict 包含 OUTPUT_COLS 中的字段，
          以及可选的 _yellow=True/False 控制是否高亮整行。
    返回 BytesIO（已 seek(0)）
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:30]

    # 写表头
    ws.append(OUTPUT_COLS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER

    # 列宽
    col_widths = {
        '日期': 12, '摘要': 36, '科目名称': 20, '科目代码': 14,
        '借方金额': 14, '贷方金额': 14, '原币金额': 14,
        '凭证号': 8, '分录序号': 8, '附件数': 6,
    }
    for i, col in enumerate(OUTPUT_COLS, 1):
        ws.column_dimensions[_col_letter(i)].width = col_widths.get(col, 10)

    # 写数据
    for row in rows:
        yellow = row.get('_yellow', False)
        values = [row.get(c, '') for c in OUTPUT_COLS]
        ws.append(values)
        if yellow:
            for cell in ws[ws.max_row]:
                cell.fill = YELLOW_FILL
        for cell in ws[ws.max_row]:
            cell.border = BORDER
            # 金额列右对齐
            if OUTPUT_COLS[cell.column - 1] in ('借方金额', '贷方金额', '原币金额', '数量', '单价', '汇率'):
                cell.alignment = Alignment(horizontal='right')

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _col_letter(n: int) -> str:
    """1-based column index to letter (A, B, ... Z, AA, ...)"""
    result = ''
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result
